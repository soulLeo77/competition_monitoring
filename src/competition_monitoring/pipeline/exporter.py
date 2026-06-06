from datetime import datetime
from pathlib import Path

from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore

from ..models.product import Product
from ..util.report import (
    COLUMN_HEADERS,
    FIELD_ORDER,
    HEADER_FILL,
    HEADER_FONT,
    ROW_ALT_FILL,
    STORE_SHEET_COL_WIDTHS,
    SUMMARY_HEADER_FILL,
    THIN_BORDER,
)


def _parse_scraped_date(s: str | None) -> str:
    if not s:
        return "N/A"
    try:
        dt = datetime.strptime(s.strip(), "%Y-%m-%d_%H-%M-%S")
        return dt.strftime("%Y-%m-%d")
    except (ValueError, AttributeError):
        return s.strip()


class ExcelExporter:
    def __init__(
        self, output_path: Path, categories: dict[str, list[str]] | None = None
    ) -> None:
        self.output_path = output_path
        self.wb = Workbook()
        self._categories = categories or {}

    def export(self, products_by_store: dict[str, list[Product]]) -> Path:
        ws_summary = self.wb.active
        assert ws_summary, "There isn't an active sheet in the workbook"
        ws_summary.title = "Summary"
        self._write_summary(ws_summary, products_by_store)

        for store, products in products_by_store.items():
            if not products:
                continue
            ws = self.wb.create_sheet(title=store[:31])
            self._write_store_sheet(ws, products)

        self.wb.save(self.output_path)
        return self.output_path

    def _write_summary(
        self, ws: Worksheet, products_by_store: dict[str, list[Product]]
    ) -> None:
        ws.title = "Summary"
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "Competition Monitoring - Summary Report"
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center")

        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")

        headers = ["Store", "Products Found", "Date Range", "Categories"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = SUMMARY_HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        total_products = 0
        row = 5
        for store, products in products_by_store.items():
            if not products:
                continue
            count = len(products)
            total_products += count
            dates = sorted(
                {
                    _parse_scraped_date(p.scraped_at)
                    for p in products
                    if _parse_scraped_date(p.scraped_at) != "N/A"
                }
            )
            date_range = (
                f"{dates[0]} - {dates[-1]}"
                if len(dates) > 1
                else (dates[0] if dates else "N/A")
            )
            cats = ", ".join(products[0].source)
            ws.cell(
                row=row, column=1, value=store.title() if store else store
            ).border = THIN_BORDER
            ws.cell(row=row, column=2, value=count).border = THIN_BORDER
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3, value=date_range).border = THIN_BORDER
            cat_cell = ws.cell(
                row=row, column=4, value=", ".join(self._get_categories(products))
            )
            cat_cell.border = THIN_BORDER
            if row % 2 == 0:
                for c in range(1, 5):
                    ws.cell(row=row, column=c).fill = ROW_ALT_FILL
            row += 1

        row += 1
        total_cell = ws.cell(row=row, column=1, value="TOTAL")
        total_cell.font = Font(name="Calibri", bold=True, size=11)
        total_cell.border = THIN_BORDER
        total_val = ws.cell(row=row, column=2, value=total_products)
        total_val.font = Font(name="Calibri", bold=True, size=11)
        total_val.alignment = Alignment(horizontal="center")
        total_val.border = THIN_BORDER
        for c in range(3, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 30

    def _get_categories(self, products: list[Product]) -> list[str]:
        if not products:
            return []
        store_key = products[0].source.lower().replace(" ", "_")
        return self._categories.get(store_key, ["laptops"])

    def _write_store_sheet(self, ws: Worksheet, products: list[Product]) -> None:
        headers = [
            COLUMN_HEADERS.get(f, f.replace("_", " ").title()) for f in FIELD_ORDER
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER

        for row_idx, product in enumerate(products, 2):
            values = [getattr(product, f) for f in FIELD_ORDER]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="center" if col_idx <= 9 else "left",
                    vertical="top",
                )
                if col_idx in (3, 4):
                    cell.alignment = Alignment(horizontal="right", vertical="top")
                if row_idx % 2 == 0:
                    cell.fill = ROW_ALT_FILL

        for col, width in STORE_SHEET_COL_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.auto_filter.ref = ws.dimensions
