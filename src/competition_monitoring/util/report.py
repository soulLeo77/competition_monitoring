from openpyxl.styles import Border, Font, PatternFill, Side  # type: ignore

FIELD_ORDER: list[str] = [
    "name",
    "brand",
    "price",
    "original_price",
    "discount",
    "currency",
    "availability",
    "rating",
    "review_count",
    "url",
    "source",
    "scraped_at",
]

COLUMN_HEADERS: dict[str, str] = {
    "name": "Product Name",
    "brand": "Brand",
    "price": "Price",
    "original_price": "Original Price",
    "discount": "Discount",
    "currency": "Currency",
    "availability": "In Stock",
    "rating": "Rating",
    "review_count": "Reviews",
    "url": "URL",
    "source": "Store",
    "scraped_at": "Scraped At",
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUMMARY_HEADER_FILL = PatternFill(
    start_color="2E75B6", end_color="2E75B6", fill_type="solid"
)
ROW_ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

STORE_SHEET_COL_WIDTHS: dict[int, int] = {
    1: 55,
    2: 18,
    3: 14,
    4: 14,
    5: 10,
    6: 10,
    7: 10,
    8: 8,
    9: 8,
    10: 50,
    11: 14,
    12: 20,
}
